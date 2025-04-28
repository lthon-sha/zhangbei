import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess

# Create a simple DataFrame
data = {
    'A': [1, 2, 3],
    'B': [4, 5, 6],
    'C': [7, 8, 9]
}
df = pd.DataFrame(data)

# Create an Excel file
wb = Workbook()
ws = wb.active

# Write DataFrame to worksheet
for r in dataframe_to_rows(df, index=False, header=True):
    ws.append(r)

# Rename column D to "Sum_A_B" and add the formula
ws['D1'] = 'Sum_A_B'
for row in range(2, len(df) + 2):
    ws[f'D{row}'] = f'=A{row} + B{row}'

# Save the Excel file
wb.save('simple_dataframe.xlsx')

# Convert Excel to PDF using LibreOffice
subprocess.run(['libreoffice', '--headless', '--convert-to', 'pdf', 'simple_dataframe.xlsx'])
